from openpyxl import *
import os

#TODO
#LIST ALL FILES IN DIRECTORY AND SUBDIRECTORY
#MANIPULATE THE DATA TO CREATE A ROW BASED ON THE FILE NAME SPLIT BY SPECIFIC PARAMETER.




# Creates an excel file and saves it using the 'myworkbook' name
filetree = "C:/Users/Brad/source/repos/File Tree to Excel/File Tree to Excel/myworkbook.xlsx"
wb = Workbook()
ws = wb.active # Gives us access to the worksheet inside
c = ws['A5']

wb.create_sheet("Here I am")
ws['A4'] = 4
c.value = 'Hello World'


wb.save(filetree) #Saves the file. This overwrites the entire file without any prompt. 


basepath = 'C:/Users/Brad/source/repos/File Tree to Excel/File Tree to Excel'

for entry in os.listdir(basepath):
    if os.path.isfile(os.path.join(basepath, entry)):
        print(entry)




